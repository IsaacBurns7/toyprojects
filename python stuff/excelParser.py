import openpyxl	
from openpyxl.utils import get_column_letter	
import pandas as pd	

#sheets: one for all, including vacant apartments	
#one for all apartment units	

custom_header = [	
    "Unit", "Bldg.", "Apt.", "First Name", "Last Name", "Gender", "Tenant Status", "Category",	
    "Lease Type", "Move In", "Move Out", "Cell Phone", "Home Phone", "Work Phone",	
    "Alternate Phone", "Personal Email Address", "Work Email Address", "Alternate Email Address"	
]	

def normalize_row(row, target_length):	
    # Ensure row has exact target_length	
    row = list(row[:target_length]) + [None] * (target_length - len(row))	

    # Normalize each cell (strip whitespace, unify empty values)	
    return tuple(	
        (str(cell).strip().lower() if isinstance(cell, str) else ("" if cell is None else str(cell)))	
        for cell in row	
    )	


def autofit_columns(sheet):	
    for col in sheet.columns:	
        max_length = 0	
        col_letter = get_column_letter(col[0].column)	

        for cell in col:	
            try:	
                cell_value = str(cell.value)	
                if cell_value:	
                    max_length = max(max_length, len(cell_value))	
            except:	
                pass	

        adjusted_width = max_length + 2  # add padding	
        sheet.column_dimensions[col_letter].width = adjusted_width	

def get_column_indices_by_headers(sheet, header_names):	
    """Return indices of specified header names."""	
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))	
    header_map = {name: idx for idx, name in enumerate(header_row)}	
    indices = []	

    for name in header_names:	
        if name in header_map:	
            indices.append(header_map[name])	
        else:	
            raise ValueError(f"Header '{name}' not found in sheet '{sheet.title}'")	

    return indices	

def collect_blank_rows(input_file, sheet_names, headers_to_check, output_sheet_name = "No Contact Info (Any)"):	
    wb = openpyxl.load_workbook(input_file)	

    if(output_sheet_name in wb.sheetnames):	
        del wb[output_sheet_name]	

    output_sheet = wb.create_sheet(output_sheet_name)	
    output_row = 1	
    seen_rows = set()	

    # Write custom header	
    for col_idx, header in enumerate(custom_header, start=1):	
        output_sheet.cell(row=output_row, column=col_idx, value=header)	
    output_row += 1	

    for sheet_name in sheet_names:	
        if sheet_name not in wb.sheetnames:	
            print(f"Sheet '{sheet_name}' not found, skipping.")	
            continue	
        sheet = wb[sheet_name]	

        try:	
            col_indices = get_column_indices_by_headers(sheet, headers_to_check)	
        except ValueError as e:	
            print(e)	
            continue	
        print(col_indices)	
        for row in sheet.iter_rows(min_row=2):	
            row_tuple = normalize_row(row, len(custom_header))	
            if row_tuple in seen_rows:	
                continue  # Skip duplicate row	

            seen_rows.add(row_tuple)	

            if any(row[idx].value is None or str(row[idx].value).strip() == "" for idx in col_indices):	
                for col_idx, cell in enumerate(row, start = 1):	
                    output_sheet.cell(row=output_row, column=col_idx, value=cell.value)	
                output_row += 1	
    autofit_columns(output_sheet)	
    wb.save(input_file)	
    print(f"Blank rows copied to '{output_sheet_name}' sheet in '{input_file}'.")	


def aggregate_sheets(input_file, sheet_names, output_sheet_name='Master Sheet'):	
    wb = openpyxl.load_workbook(input_file)	

    # Create or reset the master sheet	
    if output_sheet_name in wb.sheetnames:	
        del wb[output_sheet_name]	
    master_sheet = wb.create_sheet(output_sheet_name)	
    output_row = 1   	
    seen_rows2 = set()	
    header_written = False	

    # Write custom header	
    for col_idx, header in enumerate(custom_header, start=1):	
        master_sheet.cell(row=output_row, column=col_idx, value=header)	
    output_row += 1	

    for sheet_name in sheet_names:	
        if sheet_name not in wb.sheetnames:	
            print(f"Sheet '{sheet_name}' not found, skipping.")	
            continue	

        sheet = wb[sheet_name]	

        for i, row in enumerate(sheet.iter_rows(values_only=True)):	
            if i == 0:	
                if not header_written:	
                    for col_idx, value in enumerate(row, start=1):	
                        master_sheet.cell(row=output_row, column=col_idx, value=value)	
                    output_row += 1	
                    header_written = True	
                continue  # Skip header for other sheets	

            row_tuple = normalize_row(row, len(custom_header))	
            if row_tuple in seen_rows2:	
                continue  # Skip duplicate row	

            seen_rows2.add(row_tuple)	

    autofit_columns(master_sheet)	
    wb.save(input_file)	
    print(f"All rows aggregated into '{output_sheet_name}' in '{input_file}'.")	

if __name__ == "__main__":	
    input_excel = "tenants.xlsx"	
    sheets_to_check = ["Sheet1", "20 Bldg", "30 Bldg", ]	
    headers_to_check = [	
        "Cell Phone",	
        "Home Phone",	
        "Work Phone",	
        "Alternate Phone",	
        "Personal Email Address",	
        "Work Email Address",	
        "Alternate Email Address"	
    ]	
    #collect_blank_rows(input_excel, sheets_to_check, headers_to_check)	
    #aggregate_sheets(input_excel, sheets_to_check)